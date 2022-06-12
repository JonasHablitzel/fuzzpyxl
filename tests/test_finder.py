import fuzzpyxl
import pytest


@pytest.fixture
def DummyWorksheet():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    return ws


@pytest.fixture
def dummy_ws_with_cells(DummyWorksheet):
    ws = DummyWorksheet

    ws.cell(column=1, row=1, value="a")
    ws.cell(column=2, row=1, value="a")
    ws.cell(column=3, row=1, value="a")
    ws.cell(column=1, row=2, value="a")
    ws.cell(column=2, row=2, value="a")
    ws.cell(column=3, row=2, value="a")
    ws.cell(column=1, row=3, value="a")
    ws.cell(column=2, row=3, value="a")
    ws.cell(column=3, row=3, value="a")

    return ws


@pytest.fixture
def dummy_ws_with_differentcells(DummyWorksheet):
    ws = DummyWorksheet

    ws.cell(column=1, row=1, value="11")
    ws.cell(column=2, row=1, value="21")
    ws.cell(column=3, row=1, value="31")
    ws.cell(column=1, row=2, value="12")
    ws.cell(column=2, row=2, value="22")
    ws.cell(column=3, row=2, value="32")
    ws.cell(column=1, row=3, value="13")
    ws.cell(column=2, row=3, value="23")
    ws.cell(column=3, row=3, value="33")

    return ws


@pytest.fixture
def dummy_ws_with_doubleoccurences(DummyWorksheet):
    ws = DummyWorksheet

    ws.cell(column=1, row=1, value="11")
    ws.cell(column=2, row=1, value="11")
    ws.cell(column=3, row=1, value="11")
    ws.cell(column=1, row=2, value="22")
    ws.cell(column=2, row=2, value="22")
    ws.cell(column=3, row=2, value="22")
    ws.cell(column=1, row=3, value="33")
    ws.cell(column=2, row=3, value="33")
    ws.cell(column=3, row=3, value="33")

    return ws


def test_find_values_in_area_one(dummy_ws_with_cells):
    ws = dummy_ws_with_cells
    area = fuzzpyxl.CellArea(1, 1, 1, 1)
    found_cells = fuzzpyxl.find_values_in_area(ws, "a", area)
    assert len(found_cells) == 1


def test_find_values_in_area_all(dummy_ws_with_cells):
    ws = dummy_ws_with_cells
    area = fuzzpyxl.CellArea(1, 3, 1, 3)
    found_cells = fuzzpyxl.find_values_in_area(ws, "a", area)
    assert len(found_cells) == 9


def test_find_values_in_area_row(dummy_ws_with_cells):
    ws = dummy_ws_with_cells
    area = fuzzpyxl.CellArea(1, 3, 1, 1)
    found_cells = fuzzpyxl.find_values_in_area(ws, "a", area)
    assert len(found_cells) == 3


def test_find_values_in_area_col(dummy_ws_with_cells):
    ws = dummy_ws_with_cells
    area = fuzzpyxl.CellArea(1, 1, 1, 3)
    found_cells = fuzzpyxl.find_values_in_area(ws, "a", area)
    assert len(found_cells) == 3


def test_find_first_value_in_area_row_first_col_first_row(dummy_ws_with_cells):
    ws = dummy_ws_with_cells
    area = fuzzpyxl.CellArea(1, 3, 1, 3)
    found_cell = fuzzpyxl.find_first_value_in_area(
        ws, "a", area, "first", "first", "row"
    )
    assert found_cell.column == 1 and found_cell.row == 1


def test_find_first_value_in_area_row_last_col_first_row(dummy_ws_with_cells):
    ws = dummy_ws_with_cells
    area = fuzzpyxl.CellArea(1, 3, 1, 3)
    found_cell = fuzzpyxl.find_first_value_in_area(
        ws, "a", area, "last", "first", "row"
    )
    assert found_cell.column == 1 and found_cell.row == 3


def test_find_first_value_in_area_row_first_col_last_row(dummy_ws_with_cells):
    ws = dummy_ws_with_cells
    area = fuzzpyxl.CellArea(1, 3, 1, 3)
    found_cell = fuzzpyxl.find_first_value_in_area(
        ws, "a", area, "first", "last", "row"
    )
    assert found_cell.column == 3 and found_cell.row == 1


def test_find_first_value_in_area_row_last_col_last_row(dummy_ws_with_cells):
    ws = dummy_ws_with_cells
    area = fuzzpyxl.CellArea(1, 3, 1, 3)
    found_cell = fuzzpyxl.find_first_value_in_area(ws, "a", area, "last", "last", "row")
    assert found_cell.column == 3 and found_cell.row == 3


def test_find_first_value_in_area_row_first_col_first_col(dummy_ws_with_cells):
    ws = dummy_ws_with_cells
    area = fuzzpyxl.CellArea(1, 3, 1, 3)
    found_cell = fuzzpyxl.find_first_value_in_area(
        ws, "a", area, "first", "first", "column"
    )
    assert found_cell.column == 1 and found_cell.row == 1


def test_find_n_values_in_area_row_first_col_first_row(dummy_ws_with_cells):
    ws = dummy_ws_with_cells
    area = fuzzpyxl.CellArea(1, 3, 1, 3)
    found_cells = fuzzpyxl.find_n_values_in_area(
        ws, "a", area, "first", "first", "row", 2
    )
    assert len(found_cells) == 2
    assert found_cells[0].column == 1
    assert found_cells[0].row == 1
    assert found_cells[1].column == 2
    assert found_cells[1].row == 1


def test_find_n_values_in_area_row_first_col_first_col(dummy_ws_with_cells):
    ws = dummy_ws_with_cells
    area = fuzzpyxl.CellArea(1, 3, 1, 3)
    found_cells = fuzzpyxl.find_n_values_in_area(
        ws, "a", area, "first", "first", "column", 2
    )
    assert len(found_cells) == 2
    assert found_cells[0].column == 1
    assert found_cells[0].row == 1
    assert found_cells[1].column == 1
    assert found_cells[1].row == 2


def test_find_n_values_in_area_value_error_row_direction(dummy_ws_with_cells):
    ws = dummy_ws_with_cells
    with pytest.raises(ValueError):
        area = fuzzpyxl.CellArea(1, 3, 1, 3)
        found_cells = fuzzpyxl.find_n_values_in_area(
            ws, "a", area, "asdas", "first", "column", 2
        )


def test_find_n_values_in_area_value_error_col_direction(dummy_ws_with_cells):
    ws = dummy_ws_with_cells
    with pytest.raises(ValueError):
        area = fuzzpyxl.CellArea(1, 3, 1, 3)
        found_cells = fuzzpyxl.find_n_values_in_area(
            ws, "a", area, "first", "asd", "column", 2
        )


def test_find_n_values_in_area_value_error_major_direction(dummy_ws_with_cells):
    ws = dummy_ws_with_cells
    with pytest.raises(ValueError):
        area = fuzzpyxl.CellArea(1, 3, 1, 3)
        found_cells = fuzzpyxl.find_n_values_in_area(
            ws, "a", area, "first", "first", "asd", 2
        )


def test_find_table_area_correct(dummy_ws_with_differentcells):
    ws = dummy_ws_with_differentcells
    area = fuzzpyxl.CellArea(1, 3, 1, 3)
    res = fuzzpyxl.find_table_area(ws, area, "22", "33")
    if res is not None:
        cell1, cell2 = res
    assert cell1.column == 2 and cell1.row == 2
    assert cell2.column == 3 and cell2.row == 3


def test_find_table_area_fail_none(dummy_ws_with_doubleoccurences):
    ws = dummy_ws_with_doubleoccurences
    area = fuzzpyxl.CellArea(1, 3, 1, 3)
    res = fuzzpyxl.find_table_area(ws, area, "22", "33")
    assert res is None
